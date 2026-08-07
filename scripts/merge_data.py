"""Consolida data1/data2/data3 (exports do GAL) em planilhas prontas para import.

Os três CSVs compartilham o mesmo schema do GAL (110 colunas, `;`, ISO-8859-1),
mas NÃO são fatias disjuntas: data1∩data3 e data2∩data3 se sobrepõem, e a
granularidade é uma linha por EXAME, não por amostra (NS1, IgM e ZDC viram três
linhas do mesmo Número Interno).

Este script empilha os três, mantém só amostras de dengue (prefixo D) e colapsa
para UMA linha por amostra, preferindo a linha da PCR (ZDC). Gera dois arquivos:

    dengue_consolidado.xlsx            acervo total
    dengue_consolidado_pendentes.xlsx  sem as amostras que já fizeram PCR
                                       — é esta que vai para o sistema

Os arquivos de entrada são descobertos automaticamente: qualquer .csv/.xlsx em
data/ que tenha a coluna 'Número Interno' entra, com o nome que o GAL tiver dado.

Uso:
    python -m scripts.merge_data                    # tudo que houver em data/
    python -m scripts.merge_data a.csv b.csv        # arquivos específicos
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from src.parsing import ano_verdade, montar_chave, parse_ni

DATA = Path(__file__).resolve().parent.parent / "data"
SAIDA_TOTAL = DATA / "dengue_consolidado.xlsx"
SAIDA_PENDENTES = DATA / "dengue_consolidado_pendentes.xlsx"

# Extensões aceitas como export do GAL. O .xlsx entra porque o GAL também
# exporta nesse formato — e é o erro mais provável de quem baixar sem reparar.
_EXTENSOES = (".csv", ".xlsx", ".xls")

# Saídas do próprio script e planilhas conhecidas do projeto: nunca entram como
# ENTRADA, senão uma segunda execução consumiria o próprio resultado.
_IGNORAR = (
    SAIDA_TOTAL.name,
    SAIDA_PENDENTES.name,
    "dengue_coleta_dentro_prazo_mun_ordenado.xlsx",
)

# O importador lê a aba por nome fixo (importer.ABA) — divergir daqui faz o
# import falhar antes de ler a primeira linha.
ABA = "dengue_coleta_dentro_prazo_mun_"

COL_NI = "Número Interno"
COL_SINTOMAS = "Data do 1º Sintomas"
COL_COLETA = "Data da Coleta"
COL_EXAME = "Exame"
COL_STATUS = "Status Exame"
COL_KIT = "Kit"
COL_PROCESSAMENTO = "Data do Processamento"

EXAME_PCR = "Pesquisa de Arbovírus (ZDC)"

# Status que indicam PCR concluída no GAL.
_STATUS_CONCLUIDO = frozenset({"Resultado Liberado", "Resultado Cadastrado"})

# Janela máxima, em dias, entre o 1º sintoma e a coleta. Acima disso a carga
# viral já caiu e a PCR perde sensibilidade — a amostra não serve para
# reprocesso. É a mesma regra que gerou a planilha histórica
# "dengue_coleta_dentro_prazo_mun_ordenado" (cujo Dif Dias tem máximo 5).
MAX_DIAS_SINTOMA_COLETA = 5

# Colunas mantidas na saída. O importador usa só 6 (ver importer.COL_*), mas as
# demais dão contexto para conferência humana. Ficam de fora os campos de
# identificação pessoal (CNS, CPF, nome, endereço, telefone, nome da mãe) e os
# blocos de Sinan/Gal e tratamento/vacina — nada disso é usado pelo sistema e
# manter reduz exposição de dado sensível.
COLUNAS_SAIDA = (
    "Requisição",
    COL_NI,
    COL_SINTOMAS,
    COL_COLETA,
    "Municipio de Residência",
    "Unidade Solicitante",
    "Caso",
    COL_EXAME,
    "Metodologia",
    COL_STATUS,
    COL_KIT,
    "Fabricante",
    "Data do Processamento",
    "Data da Liberação",
    "1º Campo Resultado",
    "2º Campo Resultado",
    "3º Campo Resultado",
    "4º Campo Resultado",
    "5º Campo Resultado",
    "6º Campo Resultado",
)

# Colunas gravadas como datetime real (não texto) — ver _escrever.
COLUNAS_DATA = (COL_SINTOMAS, COL_COLETA, "Data do Processamento", "Data da Liberação")

# Prioridade de exame ao escolher a linha sobrevivente de cada amostra.
# A PCR (ZDC) vem primeiro: é o exame que o sistema de reprocesso acompanha,
# e é dele que saem os Ct do termociclador.
_PRIORIDADE_EXAME = {
    "Pesquisa de Arbovírus (ZDC)": 0,
    "Dengue, Detecção de Antígeno NS1": 1,
    "Dengue, IgM": 2,
}

# Desempate dentro do mesmo exame: uma linha com resultado liberado descreve
# melhor a amostra do que uma cancelada.
_PRIORIDADE_STATUS = {
    "Resultado Liberado": 0,
    "Resultado Cadastrado": 1,
    "Exame em Análise": 2,
    "Aguardando Triagem": 3,
    "Exame não-realizado": 4,
    "Exame Cancelado": 5,
}


def descobrir_entradas(diretorio: Path = DATA) -> list[Path]:
    """Encontra os exports do GAL num diretório.

    Aceita qualquer nome de arquivo: o que identifica um export do GAL é ter a
    coluna 'Número Interno', não se chamar data1.csv. Assim o operador larga os
    arquivos na pasta com o nome que o GAL deu e roda o script.

    As saídas do próprio script são excluídas para que reexecutar não realimente
    o resultado como entrada.
    """
    candidatos = sorted(
        p for p in diretorio.iterdir()
        if p.is_file()
        and p.suffix.lower() in _EXTENSOES
        and p.name not in _IGNORAR
        and not p.name.startswith((".", "~$"))
    )

    entradas = []
    for p in candidatos:
        try:
            colunas = _ler(p, apenas_cabecalho=True).columns
        except Exception as exc:
            print(f"    ignorado {p.name}: não foi possível ler ({type(exc).__name__})")
            continue
        if COL_NI in colunas:
            entradas.append(p)
        else:
            print(f"    ignorado {p.name}: sem coluna {COL_NI!r}")
    return entradas


def _ler(caminho: Path, *, apenas_cabecalho: bool = False) -> pd.DataFrame:
    """Lê um export do GAL preservando tudo como texto.

    dtype=str + keep_default_na=False evitam que o pandas transforme códigos em
    float (431490.0) ou vazios em NaN — ambos corromperiam identificadores.

    Aceita CSV (o formato usual, `;` e ISO-8859-1) e xlsx, porque o GAL exporta
    nos dois. Para CSV a codificação é tentada em ordem: ISO-8859-1 cobre o
    export padrão, UTF-8 cobre arquivos reprocessados por planilha.
    """
    nlinhas = 0 if apenas_cabecalho else None

    if caminho.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(caminho, dtype=str, keep_default_na=False, nrows=nlinhas)
    else:
        erro = None
        for enc in ("iso-8859-1", "utf-8-sig", "utf-8"):
            try:
                df = pd.read_csv(
                    caminho, sep=";", encoding=enc, dtype=str,
                    keep_default_na=False, nrows=nlinhas,
                )
                break
            except UnicodeDecodeError as exc:
                erro = exc
        else:
            raise erro

    df.columns = [str(c).strip() for c in df.columns]
    df["_origem"] = caminho.stem
    return df


def _data(valor):
    """Converte as datas do GAL em datetime.

    O GAL mistura dois formatos na MESMA coluna ('26-04-2026' e '26/04/26'), por
    isso os formatos são testados explicitamente, todos dia-primeiro. Deixar o
    pandas inferir leria '03/04/26' como 4 de março em vez de 3 de abril.

    Valores que já são data passam direto: a função é aplicada tanto sobre CSV
    (texto) quanto sobre xlsx já convertido, e converter duas vezes zeraria a
    coluna inteira.
    """
    if valor is None or valor is pd.NaT:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)
    if isinstance(valor, pd.Timestamp):
        return valor.to_pydatetime()

    texto = str(valor).strip()
    if not texto or texto.lower() in ("nat", "nan"):
        return None
    for fmt in ("%d/%m/%y", "%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            continue
    return None


def dias_sintoma_coleta(linha) -> Optional[int]:
    """Dias entre o 1º sintoma e a coleta, ou None se faltar alguma data.

    Usa _data() nos dois campos em vez de subtrair texto: o GAL mistura
    '26-04-2026' e '26/04/26' na mesma coluna, e ambos são dia-primeiro. Comparar
    as strings, ou deixar o pandas inferir, trocaria dia por mês em datas como
    '03/04/26' e produziria uma diferença de dias silenciosamente errada.
    """
    sintomas = _data(linha[COL_SINTOMAS])
    coleta = _data(linha[COL_COLETA])
    if sintomas is None or coleta is None:
        return None
    return (coleta - sintomas).days


def fora_do_prazo(linha) -> bool:
    """True se a coleta ocorreu tarde demais depois do 1º sintoma.

    Só exclui o que está comprovadamente ACIMA da janela. Diferença negativa
    (coleta antes do sintoma) é erro de digitação, não amostra tardia: fica no
    fluxo, sinalizada pela flag COLETA_ANTES_SINTOMA. Data ausente também não
    exclui — na dúvida a amostra permanece e a equipe decide.
    """
    dias = dias_sintoma_coleta(linha)
    return dias is not None and dias > MAX_DIAS_SINTOMA_COLETA


def ja_fez_pcr(linha) -> bool:
    """True se a PCR da amostra foi de fato CONCLUÍDA, pelos sinais do GAL.

    Exige linha de PCR (ZDC) e status concluído ou data de processamento. Os dois
    sinais concordam integralmente nos dados (601 amostras, zero divergência).

    O Kit de propósito NÃO conta como prova. Há 5 amostras com kit de PCR
    registrado e exame cancelado, sem data de processamento e sem resultado: o
    kit foi alocado, mas a PCR nunca produziu resultado — essas amostras ainda
    PRECISAM de PCR e não podem ser excluídas do fluxo da equipe.

    Kits de sorologia (ELISA/NS1) em linhas não-ZDC são irrelevantes: só a linha
    de PCR diz algo sobre a PCR.
    """
    if str(linha[COL_EXAME]).strip() != EXAME_PCR:
        return False
    return (
        str(linha[COL_STATUS]).strip() in _STATUS_CONCLUIDO
        or _data(linha[COL_PROCESSAMENTO]) is not None
    )


def resolver_ni(ni_bruto: str, data_coleta_bruta: str) -> Optional[dict]:
    """Aplica as regras da Seção 3 da ESPECIFICACAO ao NI de uma linha.

    Delega inteiramente a src.parsing — a mesma lógica que o importador usa —
    em vez de reimplementar o parse aqui. Isso traz de graça a normalização de
    caixa ('d3520/26'), o ano com zero à esquerda ('D1612/026'), o ano-de-verdade
    vindo da Data da Coleta (3.1) e a reclassificação 2026 (3.5).

    Returns:
        dict com chave/prefixo/numero/ano_verdade/flags, ou None se o NI não for
        parseável (linha descartada, como no passo 2 do importador).
    """
    p = parse_ni(ni_bruto)
    if p is None:
        return None

    coleta = _data(data_coleta_bruta)
    ano = ano_verdade(
        p.ni_ano,
        coleta,
        prefixo=p.prefixo,
        numero_sequencial=p.numero_sequencial,
    )
    if ano is None:
        return None

    return {
        "chave": montar_chave(p.prefixo, p.numero_sequencial, ano),
        "prefixo": p.prefixo,
        "numero": p.numero_sequencial,
        "ni_ano": p.ni_ano,
        "ano_verdade": ano,
        "coleta": coleta,
    }


def consolidar(df: pd.DataFrame) -> pd.DataFrame:
    """Reduz cada amostra a UMA linha, preservando o layout original do GAL.

    Não cria colunas novas: escolhe a linha mais representativa da amostra e
    descarta as demais. A prioridade é a linha de PCR (ZDC) e, dentro dela, a de
    status mais avançado — é essa que o sistema de reprocesso consome.
    """
    df = df.copy()
    df["_p_exame"] = df["Exame"].map(_PRIORIDADE_EXAME).fillna(99).astype(int)
    df["_p_status"] = df["Status Exame"].map(_PRIORIDADE_STATUS).fillna(99).astype(int)

    escolhidas = (
        df.sort_values(["_p_exame", "_p_status"], kind="stable")
        .groupby("_ni", sort=False, as_index=False)
        .head(1)
    )
    return escolhidas.drop(columns=["_p_exame", "_p_status"])


def _escrever(df: pd.DataFrame, destino: Path) -> None:
    """Grava o xlsx com datas reais e formatação legível.

    As colunas de data viram datetime de verdade (não texto): é o que a planilha
    antiga fazia e o que evita que o importador precise adivinhar dia/mês. O
    default do openpyxl seria Calibri 11 sem larguras, o que deixa a planilha com
    aparência de texto cru — aqui o corpo usa Arial (mesma fonte da planilha de
    origem do projeto), o cabeçalho fica em negrito congelado e as colunas ganham
    largura proporcional ao conteúdo.
    """
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    df = df.copy()
    for col in COLUNAS_DATA:
        if col in df.columns:
            df[col] = df[col].map(_data)

    df.to_excel(destino, index=False, engine="openpyxl", sheet_name=ABA)

    wb = load_workbook(destino)
    ws = wb.active

    corpo = Font(name="Arial", size=10)
    cabecalho = Font(name="Arial", size=10, bold=True)

    for celula in ws[1]:
        celula.font = cabecalho
        celula.alignment = Alignment(horizontal="center", vertical="center")

    colunas_data = {
        i for i, nome in enumerate(df.columns, start=1) if nome in COLUNAS_DATA
    }
    for linha in ws.iter_rows(min_row=2):
        for celula in linha:
            celula.font = corpo
            if celula.column in colunas_data:
                celula.number_format = "DD/MM/YYYY"

    # Largura pelo maior conteúdo real da coluna, com teto para não estourar a
    # tela por causa de campos de texto livre.
    for i, nome in enumerate(df.columns, start=1):
        if i in colunas_data:
            largura = max(len(str(nome)), 10) + 2
        else:
            maior = int(df[nome].astype(str).str.len().max() or 0)
            largura = max(len(str(nome)), min(maior, 40)) + 2
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(destino)


def main(argv: Optional[list[str]] = None) -> int:
    import argparse

    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("arquivos", nargs="*", type=Path,
                    help="exports do GAL (default: todos os encontrados em data/)")
    args = ap.parse_args(argv)

    if args.arquivos:
        entradas = args.arquivos
        faltando = [p for p in entradas if not p.exists()]
        if faltando:
            print(f"ERRO: não encontrado: {', '.join(str(p) for p in faltando)}")
            return 1
    else:
        print(f"[0] Procurando exports do GAL em {DATA}/")
        entradas = descobrir_entradas()

    if not entradas:
        print(f"ERRO: nenhum export do GAL encontrado em {DATA}/.\n"
              f"      Um export precisa ter a coluna {COL_NI!r}.")
        return 1

    quadros = []
    for p in entradas:
        d = _ler(p)
        quadros.append(d)
        print(f"    {p.name}: {len(d)} linhas")

    df = pd.concat(quadros, ignore_index=True)
    print(f"[1] Concatenado: {len(df)} linhas de {len(entradas)} arquivo(s)")

    faltantes = [c for c in COLUNAS_SAIDA if c not in df.columns]
    if faltantes:
        print(f"ERRO: colunas ausentes nos arquivos: {faltantes}")
        return 1

    resolvidos = [
        resolver_ni(ni, coleta)
        for ni, coleta in zip(df[COL_NI], df[COL_COLETA])
    ]
    df["_ni"] = [r["chave"] if r else "" for r in resolvidos]
    df["_prefixo"] = [r["prefixo"] if r else "" for r in resolvidos]
    df["_numero"] = [r["numero"] if r else 0 for r in resolvidos]
    df["_ano"] = [r["ano_verdade"] if r else 0 for r in resolvidos]

    # Filtro de dengue: prefixo D já normalizado em maiúsculas pelo parse_ni.
    df = df[df["_prefixo"] == "D"].copy()
    df[COL_NI] = df["_ni"]
    print(f"[2] Só amostras de dengue (prefixo D): {len(df)} linhas")

    antes = len(df)
    df = df.drop_duplicates(
        subset=[c for c in df.columns if c not in ("_origem",)]
    )
    print(f"[3] Removidas {antes - len(df)} linhas idênticas (sobreposição)")

    final = consolidar(df)
    n_pcr = (final[COL_EXAME] == EXAME_PCR).sum()
    print(f"[4] Consolidado: {len(final)} amostras únicas ({n_pcr} com linha de PCR)")

    # Ordenação canônica da Seção 3.3: ano_verdade, prefixo, numero_sequencial.
    # numero_sequencial precisa ser INTEGER (3.3) — como texto, D1000 viria
    # antes de D999.
    final["_ano"] = final["_ano"].astype(int)
    final["_numero"] = final["_numero"].astype(int)
    final = final.sort_values(
        ["_ano", "_prefixo", "_numero"], kind="stable"
    ).reset_index(drop=True)

    # Recorta para o perfil de colunas (COLUNAS_SAIDA já traz Sintomas e Coleta
    # lado a lado, logo após o Número Interno).
    final = final[list(COLUNAS_SAIDA)]

    _escrever(final, SAIDA_TOTAL)
    print(f"[5] Planilha total: {SAIDA_TOTAL.name}")
    print(f"    {len(final)} linhas x {len(final.columns)} colunas")

    # A planilha de import exclui quem já passou pela PCR e quem foi coletado
    # fora da janela de sintomas. Os filtros vêm por último, sobre o dataset já
    # consolidado e ordenado, para que os dois arquivos sejam idênticos linha a
    # linha exceto pelas removidas.
    fez_pcr = final.apply(ja_fez_pcr, axis=1)
    tardia = final.apply(fora_do_prazo, axis=1)
    pendentes = final[~fez_pcr & ~tardia].reset_index(drop=True)

    _escrever(pendentes, SAIDA_PENDENTES)
    print(f"[6] Planilha de importação: {SAIDA_PENDENTES.name}")
    print(f"    {len(pendentes)} linhas")
    print(f"    removidas: {int(fez_pcr.sum())} já fizeram PCR, "
          f"{int((tardia & ~fez_pcr).sum())} coletadas >{MAX_DIAS_SINTOMA_COLETA} "
          f"dias após o 1º sintoma")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
