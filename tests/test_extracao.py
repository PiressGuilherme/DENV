"""Testes do mapa de trabalho para extração."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
import pytest

from src import extracao


DATA_EXTRACAO = date(2026, 8, 13)


def _amostra(numero: int, *, ano: int = 2026, prefixo: str = "D"):
    return extracao.AmostraExtracao(
        chave=f"{prefixo}{numero}/{ano % 100:02d}",
        numero_sequencial=numero,
        ano_verdade=ano,
    )


def _abrir_gerado(amostras, **kwargs):
    mapa = extracao.gerar_mapa_extracao(
        amostras,
        data_extracao=DATA_EXTRACAO,
        **kwargs,
    )
    return mapa, load_workbook(
        BytesIO(mapa.conteudo), data_only=False, rich_text=True
    )


class TestRegrasDaPlaca:
    def test_posicoes_reservadas(self):
        assert len(extracao.POSICOES_AMOSTRAS) == 94
        assert extracao.POSICOES_AMOSTRAS[:9] == (
            "A2", "A3", "A4", "A5", "A6", "A7", "A8", "A9", "B2"
        )
        assert extracao.POSICOES_AMOSTRAS[-1] == "L7"
        assert extracao.POSICAO_CN == "L8"
        assert extracao.POSICAO_CP == "L9"

    def test_uma_placa_completa_tem_94_amostras_cn_e_cp(self):
        mapa, workbook = _abrir_gerado(
            [_amostra(numero) for numero in range(94, 0, -1)]
        )
        try:
            ws = workbook["Amostras"]
            assert mapa.quantidade_amostras == 94
            assert ws["A2"].value == "1/26"
            assert ws["A9"].value == "8/26"
            assert ws["B2"].value == "9/26"
            assert ws["L7"].value == "94/26"
            assert ws["L8"].value == "CN"
            assert ws["L9"].value == "CP"
        finally:
            workbook.close()

    def test_placa_parcial_mantem_vazios_e_controles_no_final(self):
        _, workbook = _abrir_gerado([_amostra(447), _amostra(12)])
        try:
            ws = workbook["Amostras"]
            assert ws["A2"].value == "12/26"
            assert ws["A3"].value == "447/26"
            assert all(ws[posicao].value is None for posicao in extracao.POSICOES_AMOSTRAS[2:])
            assert ws["L8"].value == "CN"
            assert ws["L9"].value == "CP"
        finally:
            workbook.close()

    def test_numero_interno_sem_d_e_sem_zero_artificial(self):
        assert extracao.formatar_numero_interno(_amostra(7)) == "7/26"
        assert extracao.formatar_numero_interno(_amostra(447)) == "447/26"
        assert extracao.formatar_numero_interno(_amostra(1234)) == "1234/26"


class TestValidacoes:
    def test_selecao_vazia(self):
        with pytest.raises(extracao.ErroMapaExtracao, match="ao menos uma"):
            extracao.validar_amostras([])

    def test_mais_de_94_amostras(self):
        with pytest.raises(extracao.ErroMapaExtracao, match="no máximo 94"):
            extracao.validar_amostras([_amostra(n) for n in range(1, 96)])

    def test_amostra_duplicada(self):
        with pytest.raises(extracao.ErroMapaExtracao, match="duplicadas"):
            extracao.validar_amostras([_amostra(1), _amostra(1)])

    def test_amostra_duplicada_com_zero_a_esquerda(self):
        duplicada = extracao.AmostraExtracao("D001/26", 1, 2026)
        with pytest.raises(extracao.ErroMapaExtracao, match="duplicadas"):
            extracao.validar_amostras([_amostra(1), duplicada])

    def test_chave_e_numero_precisam_ser_consistentes(self):
        inconsistente = extracao.AmostraExtracao("D2/26", 1, 2026)
        with pytest.raises(extracao.ErroMapaExtracao, match="inconsistentes"):
            extracao.validar_amostras([inconsistente])

    def test_recusa_ano_diferente_de_2026(self):
        with pytest.raises(extracao.ErroMapaExtracao, match="somente amostras de 2026"):
            extracao.validar_amostras([_amostra(1, ano=2025)])

    def test_recusa_prefixo_diferente_de_d(self):
        with pytest.raises(extracao.ErroMapaExtracao, match="Dengue válido"):
            extracao.validar_amostras([_amostra(1, prefixo="SR")])

    @pytest.mark.parametrize("placa", [0, -1, 1.5, "x", True])
    def test_numero_da_placa_precisa_ser_inteiro_positivo(self, placa):
        with pytest.raises(extracao.ErroMapaExtracao, match="inteiro positivo"):
            extracao.montar_ensaio(DATA_EXTRACAO, placa)

    def test_operador_fora_da_lista(self):
        with pytest.raises(extracao.ErroMapaExtracao, match="Operador inválido"):
            extracao.gerar_mapa_extracao(
                [_amostra(1)],
                data_extracao=DATA_EXTRACAO,
                operador="=CMD()",
            )

    def test_modelo_ausente(self, tmp_path):
        with pytest.raises(extracao.ErroMapaExtracao, match="não encontrado"):
            extracao.gerar_mapa_extracao(
                [_amostra(1)],
                data_extracao=DATA_EXTRACAO,
                caminho_modelo=tmp_path / "inexistente.xlsx",
            )


class TestMetadados:
    def test_ensaio_operador_e_nome_do_arquivo(self):
        mapa, workbook = _abrir_gerado(
            [_amostra(447)], numero_placa=2, operador="Guilherme"
        )
        try:
            assert mapa.ensaio == "DENV130826-2"
            assert mapa.nome_arquivo == (
                "DENGUE - 2 - MAPA DE TRABALHO PARA EXTRAÇÃO 13_08_2026.xlsx"
            )
            assert workbook["Amostras"]["B1"].value == "DENV130826-2"
            assert workbook["Extração"]["K16"].value == "Guilherme"
        finally:
            workbook.close()

    def test_operador_opcional_fica_vazio(self):
        _, workbook = _abrir_gerado([_amostra(1)])
        try:
            assert workbook["Extração"]["K16"].value is None
        finally:
            workbook.close()


def _assinatura_estrutural(ws) -> dict:
    formulas = {
        cell.coordinate: cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }
    estilos = {
        cell.coordinate: cell.style_id
        for row in ws.iter_rows()
        for cell in row
        if cell.has_style
    }
    validacoes = [
        (str(dv.sqref), dv.type, dv.formula1, dv.allowBlank)
        for dv in ws.data_validations.dataValidation
    ]
    colunas = {
        chave: (dim.width, dim.hidden)
        for chave, dim in ws.column_dimensions.items()
    }
    linhas = {
        chave: (dim.height, dim.hidden)
        for chave, dim in ws.row_dimensions.items()
    }
    return {
        "dimension": ws.calculate_dimension(),
        "merged": sorted(str(intervalo) for intervalo in ws.merged_cells.ranges),
        "formulas": formulas,
        "estilos": estilos,
        "validacoes": validacoes,
        "colunas": colunas,
        "linhas": linhas,
        "orientation": ws.page_setup.orientation,
        "paper_size": ws.page_setup.paperSize,
        "print_area": str(ws.print_area),
        "gridlines": ws.sheet_view.showGridLines,
    }


class TestPreservacaoDoModelo:
    def test_modelo_versionado_esta_sanitizado(self):
        workbook = load_workbook(extracao.MODELO_MAPA_EXTRACAO, data_only=False)
        try:
            ws = workbook["Amostras"]
            assert all(ws[posicao].value is None for posicao in extracao.POSICOES_AMOSTRAS)
            assert ws["L8"].value == "CN"
            assert ws["L9"].value == "CP"
            assert ws["B1"].value is None
            assert workbook["Extração"]["K16"].value is None
        finally:
            workbook.close()

    def test_geracao_preserva_estrutura_formulas_e_estilos(self):
        modelo = load_workbook(
            extracao.MODELO_MAPA_EXTRACAO, data_only=False, rich_text=True
        )
        _, gerado = _abrir_gerado([_amostra(447)], operador="Anna")
        try:
            assert gerado.sheetnames == modelo.sheetnames
            for nome in modelo.sheetnames:
                assert _assinatura_estrutural(gerado[nome]) == _assinatura_estrutural(
                    modelo[nome]
                )
            assert gerado["Extração"]["C6"].value == "=Amostras!A2"
            assert gerado["Extração"]["N13"].value == "=Amostras!L9"
            assert gerado["Extração"]["L4"].value == '=concat("Ensaio: ",Amostras!B1)'
            assert gerado.calculation.fullCalcOnLoad is True
            assert gerado.calculation.forceFullCalc is True
        finally:
            modelo.close()
            gerado.close()

    def test_geracao_preserva_tamanhos_de_fonte_do_texto_rico(self):
        _, gerado = _abrir_gerado([_amostra(447)])
        try:
            formulario = gerado["Extração"]["D2"].value
            pagina = gerado["Extração"]["M1"].value
            assert isinstance(formulario, CellRichText)
            assert isinstance(pagina, CellRichText)

            runs_formulario = [
                (parte.text, parte.font.sz, parte.font.b)
                for parte in formulario
                if isinstance(parte, TextBlock)
            ]
            runs_pagina = [
                (parte.text, parte.font.sz, parte.font.b)
                for parte in pagina
                if isinstance(parte, TextBlock)
            ]
            assert runs_formulario == [
                ("FORMULÁRIO", 8.0, False),
                ("\n", 8.0, True),
                ("LACEN/CEVS", 10.0, True),
                ("\n", 12.0, True),
            ]
            assert runs_pagina == [
                ("REVISÃO", 12.0, False),
                ("\n00\n\n", 12.0, True),
                ("Página 1 de 1", 7.0, False),
            ]
        finally:
            gerado.close()

    def test_dropdowns_de_kit_e_operador_permanecem_no_xlsx(self):
        _, gerado = _abrir_gerado([_amostra(447)])
        try:
            validacoes = {
                str(dv.sqref): dv
                for dv in gerado["Extração"].data_validations.dataValidation
            }
            assert validacoes["C4"].type == "list"
            assert "Loccus" in validacoes["C4"].formula1
            assert validacoes["K16"].type == "list"
            assert "Guilherme" in validacoes["K16"].formula1
        finally:
            gerado.close()

    def test_modelo_faz_parte_da_arvore_data_copiada_pelo_docker(self):
        assert extracao.MODELO_MAPA_EXTRACAO.is_file()
        assert extracao.MODELO_MAPA_EXTRACAO.parent.name == "modelos"
        assert extracao.MODELO_MAPA_EXTRACAO.parents[1].name == "data"
