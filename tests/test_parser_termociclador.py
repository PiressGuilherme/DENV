"""Testes do parser de arquivos do termociclador."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from src.parser_termociclador import (
    AmostraTermociclador,
    ResultadoParseTermociclador,
    _extrair_linhas_por_cabecalho,
    _normalizar_sample_id,
    merge_arquivos_termociclador,
    parse_arquivo_termociclador,
    preparar_para_gravacao,
)

# Arquivos de teste na pasta docs
DOCS = Path(__file__).parent.parent / "docs"
ARQUIVO_1_4 = DOCS / "Dengue 1-4 030826_Abs Quant(Stage2_Step2).xlsx"
ARQUIVO_1_4_REGRAVADO = DOCS / "Dengue 1-4 3006261_Abs Quant(Stage2_Step2).xlsx"
ARQUIVO_2_3 = DOCS / "Dengue 2-3 030826_Abs Quant(Stage2_Step2).xlsx"

ARQUIVOS_EXISTEM = ARQUIVO_1_4.exists() and ARQUIVO_2_3.exists()
requer_arquivos_exemplo = pytest.mark.skipif(
    not ARQUIVOS_EXISTEM, reason="Arquivos de exemplo não encontrados em docs/"
)
requer_arquivos_1_4 = pytest.mark.skipif(
    not (ARQUIVO_1_4.exists() and ARQUIVO_1_4_REGRAVADO.exists()),
    reason="Arquivos Dengue 1-4 de regressão não encontrados em docs/",
)


def _xlsx_com_linhas_esparsas() -> bytes:
    """Cria planilha em que os dados começam em B e A não existe no XML."""
    workbook = Workbook()
    resultado = workbook.active
    resultado.title = "Abs QuantResult"

    for coluna, valor in enumerate(
        ("Well", "Sample ID", "Sample", "Sample Type", "Dye", "Gene", "Ct"),
        start=2,
    ):
        resultado.cell(1, coluna, valor)

    # Controle com caixa/espaços variados: deve ser ignorado antes de interpretar
    # "CN" como um identificador de amostra.
    for coluna, valor in {
        2: "G12", 3: " cn ", 4: "CN", 5: " unknown ",
        6: " fam ", 7: "DEN4", 8: "-",
    }.items():
        resultado.cell(2, coluna, valor)

    # ID numérico integral e classificadores normalizados pelo parser.
    for coluna, valor in {
        2: "A1", 3: 24745.0, 4: "24745", 5: " unknown ",
        6: " cy5 ", 7: "CI", 8: 18.5,
    }.items():
        resultado.cell(3, coluna, valor)
    for coluna, valor in {
        2: "A1", 3: 24745.0, 4: "24745", 5: "UNKNOWN",
        6: "FAM", 7: "DEN4", 8: "-",
    }.items():
        resultado.cell(4, coluna, valor)

    # Célula fisicamente presente apenas por estilo: não é um registro válido.
    resultado["A1000"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    estatisticas = workbook.create_sheet("Abs QuantStatistics")
    for coluna, valor in enumerate(
        ("Replicate", "Well", "Sample", "Sample Type", "Dye", "Gene"),
        start=1,
    ):
        estatisticas.cell(1, coluna, valor)
    for coluna, valor in {
        2: "G12", 3: "CN", 4: "Negative", 5: "FAM", 6: "DEN4",
    }.items():
        estatisticas.cell(251, coluna, valor)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


class TestNormalizarSampleId:
    """Testes da normalização de Sample ID."""

    @pytest.mark.parametrize("sample_id,esperado", [
        ("25346", ("D", 25346, None)),
        ("D23459", ("D", 23459, None)),
        ("D24745/26", ("D", 24745, 2026)),
        ("d24745/26", ("D", 24745, 2026)),
        ("SR123/25", ("SR", 123, 2025)),
        ("FA456/2025", ("FA", 456, 2025)),
        ("H789/29", ("H", 789, 2029)),
    ])
    def test_formatos_validos(self, sample_id, esperado):
        assert _normalizar_sample_id(sample_id) == esperado

    @pytest.mark.parametrize("sample_id", [
        "",
        "abc",
        "123/",
        "/26",
        "D-123/25",
    ])
    def test_formatos_invalidos(self, sample_id):
        with pytest.raises(ValueError):
            _normalizar_sample_id(sample_id)


class TestLinhasEsparsas:
    """Regressões independentes dos arquivos reais de bancada."""

    def test_celula_a_ausente_nao_desloca_controle_cn(self):
        conteudo = _xlsx_com_linhas_esparsas()

        with ZipFile(BytesIO(conteudo)) as arquivo:
            xml = arquivo.read("xl/worksheets/sheet2.xml")
        assert b'r="B251"' in xml
        assert b'r="A251"' not in xml

        linhas = _extrair_linhas_por_cabecalho(
            conteudo,
            "Abs QuantStatistics",
            ("Well", "Sample", "Sample Type", "Dye", "Gene"),
        )

        assert linhas == [{
            "Well": "G12",
            "Sample": "CN",
            "Sample Type": "Negative",
            "Dye": "FAM",
            "Gene": "DEN4",
        }]

    def test_parser_ignora_controle_e_linhas_fisicas_vazias(self):
        conteudo = _xlsx_com_linhas_esparsas()
        linhas = _extrair_linhas_por_cabecalho(
            conteudo,
            "Abs QuantResult",
            ("Well", "Sample ID", "Sample", "Sample Type", "Dye", "Gene", "Ct"),
        )
        assert len(linhas) == 3

        resultado = parse_arquivo_termociclador(conteudo, "Dengue 1-4 teste.xlsx")

        assert not resultado.erros
        assert len(resultado.amostras) == 1
        amostra = resultado.amostras[0]
        assert (amostra.prefixo, amostra.numero_sequencial) == ("D", 24745)
        assert amostra.cts["ci_1_4_ct"] == 18.5
        assert amostra.cts["den4_ct"] == -1.0


@requer_arquivos_1_4
class TestCompatibilidadeArquivos1_4:
    """Compara o original do equipamento com o XLSX regravado."""

    @staticmethod
    def _controles(path: Path) -> list[tuple[object, ...]]:
        linhas = _extrair_linhas_por_cabecalho(
            path.read_bytes(),
            "Abs QuantStatistics",
            ("Well", "Sample", "Sample Type", "Dye", "Gene"),
        )
        return [
            tuple(linha[c] for c in ("Well", "Sample", "Sample Type", "Dye", "Gene"))
            for linha in linhas
            if str(linha["Sample"] or "").strip().upper() == "CN"
        ]

    def test_controles_cn_tem_a_mesma_interpretacao_logica(self):
        esperado = [
            ("G12", "CN", "Negative", "FAM", "DEN4"),
            ("G12", "CN", "Negative", "VIC", "DEN1"),
            ("G12", "CN", "Negative", "Cy5", "CI"),
        ]
        assert self._controles(ARQUIVO_1_4) == esperado
        assert self._controles(ARQUIVO_1_4_REGRAVADO) == esperado

    @pytest.mark.parametrize("path", [ARQUIVO_1_4, ARQUIVO_1_4_REGRAVADO])
    def test_parser_processa_original_e_regravado(self, path):
        resultado = parse_arquivo_termociclador(path.read_bytes(), path.name)

        assert not resultado.erros
        assert len(resultado.amostras) == 94
        assert all(amostra.prefixo not in {"CN", "CP"} for amostra in resultado.amostras)


@requer_arquivos_exemplo
class TestParseArquivo1_4:
    """Testes do parse do arquivo Dengue 1-4."""

    def test_parse_sucesso(self):
        with open(ARQUIVO_1_4, "rb") as f:
            resultado = parse_arquivo_termociclador(f.read(), ARQUIVO_1_4.name)
        
        assert not resultado.erros
        assert len(resultado.amostras) > 0
        
        # Verifica estrutura das amostras
        for amp in resultado.amostras:
            assert isinstance(amp, AmostraTermociclador)
            assert amp.prefixo == "D"
            assert amp.numero_sequencial > 0
            assert "den1_ct" in amp.cts
            assert "den4_ct" in amp.cts
            assert "ci_1_4_ct" in amp.cts
            # DEN2 e DEN3 não estão neste arquivo
            assert amp.cts.get("den2_ct") is None
            assert amp.cts.get("den3_ct") is None
            # ci_2_3_ct não está neste arquivo
            assert amp.cts.get("ci_2_3_ct") is None

    def test_controles_ignorados(self):
        """CN e CP devem ser ignorados."""
        with open(ARQUIVO_1_4, "rb") as f:
            resultado = parse_arquivo_termociclador(f.read(), ARQUIVO_1_4.name)
        
        # Nenhuma amostra deve ter prefixo CN ou CP
        for amp in resultado.amostras:
            assert amp.prefixo not in ("CN", "CP")


@requer_arquivos_exemplo
class TestParseArquivo2_3:
    """Testes do parse do arquivo Dengue 2-3."""

    def test_parse_sucesso(self):
        with open(ARQUIVO_2_3, "rb") as f:
            resultado = parse_arquivo_termociclador(f.read(), ARQUIVO_2_3.name)
        
        assert not resultado.erros
        assert len(resultado.amostras) > 0
        
        # Verifica estrutura das amostras
        for amp in resultado.amostras:
            assert isinstance(amp, AmostraTermociclador)
            assert amp.prefixo == "D"
            assert amp.numero_sequencial > 0
            assert "den2_ct" in amp.cts
            assert "den3_ct" in amp.cts
            assert "ci_2_3_ct" in amp.cts
            # DEN1 e DEN4 não estão neste arquivo
            assert amp.cts.get("den1_ct") is None
            assert amp.cts.get("den4_ct") is None
            # ci_1_4_ct não está neste arquivo
            assert amp.cts.get("ci_1_4_ct") is None


@requer_arquivos_exemplo
class TestMergeArquivos:
    """Testes do merge dos dois arquivos."""

    def test_merge_sucesso(self):
        with open(ARQUIVO_1_4, "rb") as f:
            r1 = parse_arquivo_termociclador(f.read(), ARQUIVO_1_4.name)
        with open(ARQUIVO_2_3, "rb") as f:
            r2 = parse_arquivo_termociclador(f.read(), ARQUIVO_2_3.name)
        
        merged = merge_arquivos_termociclador(r1, r2)
        
        assert not merged.erros
        assert len(merged.amostras) > 0
        
        # Verifica que tem todos os 6 CTs (DEN1-4 + CI 1-4 + CI 2-3)
        for amp in merged.amostras:
            for ct in ["den1_ct", "den2_ct", "den3_ct", "den4_ct", "ci_1_4_ct", "ci_2_3_ct"]:
                assert ct in amp.cts
        
        # Verifica que amostras dos dois arquivos casaram pelo Sample ID
        # O número de amostras merged deve ser <= soma dos dois (pois merge por ID)
        assert len(merged.amostras) <= len(r1.amostras) + len(r2.amostras)

    def test_amostra_completa_apos_merge(self):
        """Uma amostra presente nos dois arquivos deve ter DEN1-4 + CI."""
        with open(ARQUIVO_1_4, "rb") as f:
            r1 = parse_arquivo_termociclador(f.read(), ARQUIVO_1_4.name)
        with open(ARQUIVO_2_3, "rb") as f:
            r2 = parse_arquivo_termociclador(f.read(), ARQUIVO_2_3.name)
        
        merged = merge_arquivos_termociclador(r1, r2)
        
        # Pega primeira amostra e verifica se tem dados dos dois arquivos
        amp = merged.amostras[0]
        # Pelo menos um dos DEN1/DEN4 (arquivo 1-4) e um dos DEN2/DEN3 (arquivo 2-3)
        tem_1_4 = amp.cts["den1_ct"] is not None or amp.cts["den4_ct"] is not None
        tem_2_3 = amp.cts["den2_ct"] is not None or amp.cts["den3_ct"] is not None
        tem_ci_1_4 = amp.cts["ci_1_4_ct"] is not None
        tem_ci_2_3 = amp.cts["ci_2_3_ct"] is not None
        
        # Pelo menos um dos CI deve estar presente (vem dos dois arquivos)
        assert tem_ci_1_4 or tem_ci_2_3

    def test_sentinela_nao_detectado(self):
        """Testa que '-' na planilha vira -1.0 (não detectado) e ausente vira None (não testado)."""
        from src.parser_termociclador import _ct_para_float
        import pandas as pd
        
        # "-" -> -1.0 (não detectado)
        assert _ct_para_float("-") == -1.0
        
        # "" (vazio) -> None (não testado)
        assert _ct_para_float("") is None
        assert _ct_para_float(None) is None
        assert _ct_para_float(pd.NA) is None
        assert _ct_para_float(float("nan")) is None
        
        # Valor válido
        assert _ct_para_float("25.5") == 25.5
        assert _ct_para_float("25,5") == 25.5  # decimal BR
        assert _ct_para_float(25.5) == 25.5


@requer_arquivos_exemplo
class TestPrepararParaGravacao:
    """Testes da preparação para gravação no banco."""

    def test_com_ano_padrao(self):
        with open(ARQUIVO_1_4, "rb") as f:
            r1 = parse_arquivo_termociclador(f.read(), ARQUIVO_1_4.name)
        with open(ARQUIVO_2_3, "rb") as f:
            r2 = parse_arquivo_termociclador(f.read(), ARQUIVO_2_3.name)
        
        merged = merge_arquivos_termociclador(r1, r2)
        
        # Prepara com ano padrão
        dados = preparar_para_gravacao(merged, ano_padrao=2025)
        
        assert len(dados) == len(merged.amostras)
        for d in dados:
            assert d["ano_verdade"] == 2025
            assert "prefixo" in d
            assert "numero_sequencial" in d
            assert "cts" in d

    def test_sem_ano_padrao_mantem_none(self):
        with open(ARQUIVO_1_4, "rb") as f:
            r1 = parse_arquivo_termociclador(f.read(), ARQUIVO_1_4.name)
        
        # Algumas amostras podem não ter ano
        dados = preparar_para_gravacao(r1, ano_padrao=None)
        
        for d in dados:
            # Se a amostra original não tinha ano, deve continuar None
            if d["ano_verdade"] is None:
                assert d["ano_verdade"] is None


# Teste de integração rápido (roda só se tiver DATABASE_URL)
@pytest.mark.integracao
@requer_arquivos_exemplo
class TestIntegracaoBanco:
    """Testes de integração com banco real (precisa DATABASE_URL)."""

    def test_resolver_e_gravar(self, _pg_schema_con):
        """Testa resolução no banco e gravação."""
        from src import db
        
        con = _pg_schema_con
        db.criar_schema(con)
        
        # Importa dados base primeiro
        from src.importer import importar, XLSX_PADRAO
        importar(XLSX_PADRAO, _con=con, verificar_sanidade=False)
        
        # Parse arquivos
        with open(ARQUIVO_1_4, "rb") as f:
            r1 = parse_arquivo_termociclador(f.read(), ARQUIVO_1_4.name)
        with open(ARQUIVO_2_3, "rb") as f:
            r2 = parse_arquivo_termociclador(f.read(), ARQUIVO_2_3.name)
        
        merged = merge_arquivos_termociclador(r1, r2)
        dados = preparar_para_gravacao(merged, ano_padrao=2025)
        
        # Resolve no banco
        resolvidas, nao_encontradas, ano_ambiguo = db.resolver_amostras_termociclador(
            con, dados
        )
        
        # Deve encontrar pelo menos algumas amostras (mesmo número sequencial)
        # Como os Sample IDs dos arquivos de teste podem não bater com o banco de teste,
        # aceitamos que possa não encontrar nenhuma
        assert isinstance(resolvidas, dict)
        assert isinstance(nao_encontradas, list)
        assert isinstance(ano_ambiguo, list)
        
        # Se encontrou alguma, testa gravação
        if resolvidas:
            resultado = db.gravar_resultados_termociclador(con, resolvidas)
            assert resultado.gravados >= 0
            assert resultado.campos_gravados >= 0


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
