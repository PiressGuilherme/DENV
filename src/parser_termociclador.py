"""Parser para arquivos de resultado do termociclador (Abs Quant).

Formato: aba 'Abs QuantResult' com colunas:
    Well, Sample ID, Sample, Sample Type, Dye, Gene, Test Name, Ct, ...

Dois arquivos por corrida:
    - Dengue 1-4: FAM=DEN4, VIC=DEN1, Cy5=CI (-> ci_1_4_ct)
    - Dengue 2-3: FAM=DEN2, VIC=DEN3, Cy5=CI (-> ci_2_3_ct)

Sample ID pode vir em 3 formatos:
    - 25346          -> prefixo="D", numero=25346, ano=None
    - D23459         -> prefixo="D", numero=23459, ano=None
    - D24745/26      -> prefixo="D", numero=24745, ano=2026

Amostras com Sample Type != "Unknown" (CN, CP) são ignoradas.

Valores de Ct:
    - None = não testado (sorotipo não presente no arquivo / coluna ausente)
    - -1.0 = não detectado (veio "-" na planilha)
    - >0 = valor de Ct válido
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


@dataclass
class AmostraTermociclador:
    """Uma amostra parseada do termociclador."""
    prefixo: str
    numero_sequencial: int
    ano_verdade: Optional[int]  # None se não veio no Sample ID
    # den1_ct..den4_ct, ci_1_4_ct, ci_2_3_ct
    # None = não testado, -1.0 = não detectado, >0 = Ct válido
    cts: dict[str, Optional[float]] = field(default_factory=dict)


@dataclass
class ResultadoParseTermociclador:
    """Resultado do parse dos arquivos do termociclador."""
    amostras: list[AmostraTermociclador] = field(default_factory=list)
    sample_ids_sem_ano: list[str] = field(default_factory=list)  # para pedir ano na UI
    erros: list[str] = field(default_factory=list)


# Mapeamento Dye -> Sorotipo por arquivo
# CI é mapeado para coluna específica por arquivo (ci_1_4_ct ou ci_2_3_ct)
MAPA_DYE_SOROTIPO = {
    "Dengue 1-4": {
        "FAM": "DEN4",
        "VIC": "DEN1",
        "CY5": "CI_1_4",  # especial: vira ci_1_4_ct
    },
    "Dengue 2-3": {
        "FAM": "DEN2",
        "VIC": "DEN3",
        "CY5": "CI_2_3",  # especial: vira ci_2_3_ct
    },
}

CABECALHOS_RESULTADO = (
    "Well",
    "Sample ID",
    "Sample",
    "Sample Type",
    "Dye",
    "Gene",
    "Ct",
)

# Colunas de Ct do banco (DEN1-4)
COLUNAS_CT_ESPERADAS = {"den1_ct", "den2_ct", "den3_ct", "den4_ct"}
# Colunas de CI (uma por arquivo)
COLUNAS_CI_ESPERADAS = {"ci_1_4_ct", "ci_2_3_ct"}
# Todas as colunas esperadas no resultado final
TODAS_COLUNAS_ESPERADAS = COLUNAS_CT_ESPERADAS | COLUNAS_CI_ESPERADAS
AMOSTRAS_CONTROLE = {"CN", "CP"}


def _valor_ausente(valor) -> bool:
    """Retorna True para None/NaN/pd.NA sem confundir zero com vazio."""
    if valor is None:
        return True
    try:
        return bool(pd.isna(valor))
    except (TypeError, ValueError):
        return False


def _normalizar_texto(valor, *, upper: bool = False) -> str:
    """Normaliza texto usado em comparações, preservando valores numéricos."""
    if _valor_ausente(valor):
        return ""
    texto = str(valor).strip()
    return texto.upper() if upper else texto


def _normalizar_cabecalho(valor) -> str:
    """Normaliza caixa e espaços de um cabeçalho sem depender de estilo."""
    return re.sub(r"\s+", " ", _normalizar_texto(valor)).casefold()


def _extrair_linhas_por_cabecalho(
    conteudo: bytes,
    nome_aba: str,
    cabecalhos: tuple[str, ...],
) -> list[dict[str, object]]:
    """Extrai linhas usando a coluna real de cada cabeçalho.

    O openpyxl materializa células por suas coordenadas (por exemplo, ``B251``),
    portanto uma célula ``A251`` ausente no XML continua sendo logicamente vazia
    e não desloca as demais. O cabeçalho pode estar em qualquer linha; linhas sem
    nenhum valor nos campos solicitados são descartadas.
    """
    workbook = load_workbook(BytesIO(conteudo), read_only=False, data_only=True)
    try:
        if nome_aba not in workbook.sheetnames:
            raise ValueError(f"Aba '{nome_aba}' não encontrada na planilha")

        sheet = workbook[nome_aba]
        esperados = {_normalizar_cabecalho(nome): nome for nome in cabecalhos}
        colunas: dict[str, int] = {}
        melhor_conjunto: set[str] = set()
        linha_cabecalho: Optional[int] = None

        # No modo normal, max_row/max_column são calculados a partir das células
        # efetivamente carregadas, e não do <dimension> declarado pelo produtor.
        for row in sheet.iter_rows():
            encontrados: dict[str, int] = {}
            for cell in row:
                nome = esperados.get(_normalizar_cabecalho(cell.value))
                if nome is not None:
                    encontrados[nome] = cell.column
            if len(encontrados) > len(melhor_conjunto):
                melhor_conjunto = set(encontrados)
            if set(cabecalhos).issubset(encontrados):
                colunas = encontrados
                linha_cabecalho = row[0].row
                break

        if linha_cabecalho is None:
            faltando = [nome for nome in cabecalhos if nome not in melhor_conjunto]
            raise ValueError(
                "Colunas faltando na planilha: " + ", ".join(faltando)
            )

        linhas: list[dict[str, object]] = []
        for row in sheet.iter_rows(min_row=linha_cabecalho + 1):
            # O dicionário é indexado pelo número real da coluna da célula. Não
            # usamos a posição da célula dentro da sequência XML.
            valores_por_coluna = {cell.column: cell.value for cell in row}
            valores = {
                nome: valores_por_coluna.get(numero_coluna)
                for nome, numero_coluna in colunas.items()
            }
            if all(_valor_ausente(valor) for valor in valores.values()):
                continue
            linhas.append(valores)

        return linhas
    finally:
        workbook.close()


def _normalizar_sample_id(sample_id: str) -> tuple[str, int, Optional[int]]:
    """
    Normaliza Sample ID para (prefixo, numero_sequencial, ano_verdade).
    
    Formatos aceitos:
        - 25346          -> ("D", 25346, None)
        - D23459         -> ("D", 23459, None)
        - D24745/26      -> ("D", 24745, 2026)
        - d24745/26      -> ("D", 24745, 2026)  (case insensitive)
        - SR123/25       -> ("SR", 123, 2025)
    """
    if not sample_id or not isinstance(sample_id, str):
        raise ValueError(f"Sample ID inválido: {sample_id!r}")
    
    sample_id = sample_id.strip()
    
    # Regex: prefixo opcional (letras), número, /ano opcional
    # Exemplos: "25346", "D23459", "D24745/26", "SR123/25"
    m = re.match(r'^([A-Za-z]*)(\d+)(?:/(\d+))?$', sample_id)
    if not m:
        raise ValueError(f"Sample ID não parseável: {sample_id!r}")
    
    prefixo_raw, numero_raw, ano_raw = m.groups()
    prefixo = (prefixo_raw or "D").upper()
    numero = int(numero_raw)
    
    ano = None
    if ano_raw:
        # Normaliza ano: 2 dígitos -> 2000+, 4 dígitos -> como está
        if len(ano_raw) <= 2:
            ano = 2000 + int(ano_raw)
        else:
            ano = int(ano_raw)
    
    return prefixo, numero, ano


def _ct_para_float(valor) -> Optional[float]:
    """Converte valor de Ct para float.
    
    Retorna:
        - None: não testado (valor ausente/NaN)
        - -1.0: não detectado (veio "-" na planilha)
        - float > 0: valor de Ct válido
    """
    if valor is None:
        return None
    if isinstance(valor, float) and pd.isna(valor):
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass
    
    s = str(valor).strip()
    if not s:
        return None
    if s == "-":
        return -1.0  # sentinela: não detectado
    
    # Substitui vírgula por ponto (decimal BR)
    s = s.replace(",", ".")
    
    try:
        return float(s)
    except ValueError:
        return None


def _detectar_tipo_arquivo(nome_arquivo: str) -> Optional[str]:
    """Detecta se é arquivo 'Dengue 1-4' ou 'Dengue 2-3' pelo nome."""
    nome_lower = nome_arquivo.lower()
    if "1-4" in nome_lower or "1_4" in nome_lower:
        return "Dengue 1-4"
    if "2-3" in nome_lower or "2_3" in nome_lower:
        return "Dengue 2-3"
    return None


def parse_arquivo_termociclador(
    conteudo: bytes,
    nome_arquivo: str,
) -> ResultadoParseTermociclador:
    """
    Parseia um arquivo .xlsx do termociclador (aba Abs QuantResult).
    
    Returns:
        ResultadoParseTermociclador com amostras, sample_ids_sem_ano e erros.
    """
    tipo = _detectar_tipo_arquivo(nome_arquivo)
    if not tipo:
        return ResultadoParseTermociclador(
            erros=[f"Não foi possível identificar tipo do arquivo: {nome_arquivo}. "
                   "Nome deve conter '1-4' ou '2-3'."]
        )
    
    mapa_dye = MAPA_DYE_SOROTIPO[tipo]
    
    try:
        linhas = _extrair_linhas_por_cabecalho(
            conteudo,
            "Abs QuantResult",
            CABECALHOS_RESULTADO,
        )
    except Exception as e:
        return ResultadoParseTermociclador(
            erros=[f"Erro ao ler aba 'Abs QuantResult': {e}"]
        )
    
    # Agrupa por Sample ID
    amostras_dict: dict[tuple[str, int, Optional[int]], AmostraTermociclador] = {}
    sample_ids_sem_ano_set: set[str] = set()
    
    for row in linhas:
        sample_id_raw = row.get("Sample ID")
        sample = _normalizar_texto(row.get("Sample"), upper=True)
        sample_type = _normalizar_texto(row.get("Sample Type"), upper=True)
        dye = _normalizar_texto(row.get("Dye"), upper=True)
        gene = _normalizar_texto(row.get("Gene"), upper=True)
        ct_raw = row.get("Ct")
        
        # Ignora controles (CN, CP) e linhas sem Sample ID
        if _valor_ausente(sample_id_raw) or not _normalizar_texto(sample_id_raw):
            continue
        sample_id_classificacao = _normalizar_texto(sample_id_raw, upper=True)
        if sample in AMOSTRAS_CONTROLE or sample_id_classificacao in AMOSTRAS_CONTROLE:
            continue
        if sample_type != "UNKNOWN":
            continue
        
        # Normaliza Sample ID
        try:
            sample_id = _normalizar_texto(sample_id_raw)
            if isinstance(sample_id_raw, float) and sample_id_raw.is_integer():
                sample_id = str(int(sample_id_raw))
            prefixo, numero, ano = _normalizar_sample_id(sample_id)
        except ValueError as e:
            return ResultadoParseTermociclador(
                erros=[f"Sample ID inválido na linha: {sample_id_raw} - {e}"]
            )
        
        # Mapeia Dye para sorotipo
        sorotipo = mapa_dye.get(dye)
        if not sorotipo:
            continue  # Dye não mapeado (ex: ROX)
        
        # Converte Ct
        ct_valor = _ct_para_float(ct_raw)
        
        # Valida faixa de Ct
        if ct_valor is not None:
            from src.db import CT_MIN, CT_MAX
            if not (CT_MIN <= ct_valor <= CT_MAX):
                ct_valor = None  # Fora da faixa plausível -> trata como não detectado
        
        # Chave de agrupamento
        key = (prefixo, numero, ano)
        
        if key not in amostras_dict:
            amostras_dict[key] = AmostraTermociclador(
                prefixo=prefixo,
                numero_sequencial=numero,
                ano_verdade=ano,
                cts={},
            )
            if ano is None:
                sample_ids_sem_ano_set.add(f"{prefixo}{numero}")
        
        # Armazena Ct na coluna correspondente
        coluna_ct = f"{sorotipo.lower()}_ct"
        amostras_dict[key].cts[coluna_ct] = ct_valor
    
    # Converte para lista
    amostras = list(amostras_dict.values())
    
    # Garante que todas as colunas de Ct existam (None se ausente)
    for amp in amostras:
        for col in COLUNAS_CT_ESPERADAS:
            amp.cts.setdefault(col, None)
    
    return ResultadoParseTermociclador(
        amostras=amostras,
        sample_ids_sem_ano=sorted(sample_ids_sem_ano_set),
    )


def merge_arquivos_termociclador(
    resultado_1_4: ResultadoParseTermociclador,
    resultado_2_3: ResultadoParseTermociclador,
) -> ResultadoParseTermociclador:
    """
    Faz merge dos resultados dos dois arquivos (1-4 e 2-3).
    
    Une por (prefixo, numero_sequencial, ano_verdade).
    Se houver conflito de Ct no mesmo campo, mantém o não-None (ou o primeiro).
    Garante que todas as 5 colunas de CT existam no resultado.
    """
    # Índice por chave
    merged: dict[tuple[str, int, Optional[int]], AmostraTermociclador] = {}
    
    for resultado in (resultado_1_4, resultado_2_3):
        for amp in resultado.amostras:
            key = (amp.prefixo, amp.numero_sequencial, amp.ano_verdade)
            if key not in merged:
                merged[key] = AmostraTermociclador(
                    prefixo=amp.prefixo,
                    numero_sequencial=amp.numero_sequencial,
                    ano_verdade=amp.ano_verdade,
                    cts={},
                )
            # Merge dos CTs: se já tem valor, mantém; se não tem, pega o novo
            for col, val in amp.cts.items():
                if val is not None and merged[key].cts.get(col) is None:
                    merged[key].cts[col] = val
    
    # Garante que todas as colunas de CT e CI existam em todas as amostras merged
    for amp in merged.values():
        for col in COLUNAS_CT_ESPERADAS:
            amp.cts.setdefault(col, None)
        for col in COLUNAS_CI_ESPERADAS:
            amp.cts.setdefault(col, None)
    
    # Une sample_ids_sem_ano
    sem_ano = set(resultado_1_4.sample_ids_sem_ano) | set(resultado_2_3.sample_ids_sem_ano)
    
    # Une erros
    erros = resultado_1_4.erros + resultado_2_3.erros
    
    return ResultadoParseTermociclador(
        amostras=list(merged.values()),
        sample_ids_sem_ano=sorted(sem_ano),
        erros=erros,
    )


def preparar_para_gravacao(
    resultado: ResultadoParseTermociclador,
    ano_padrao: Optional[int] = None,
) -> list[dict]:
    """
    Prepara lista de dicts para db.gravar_resultados_termociclador.
    
    Se ano_padrao for fornecido, aplica às amostras sem ano.
    Retorna lista de dicts com: prefixo, numero_sequencial, ano_verdade, cts
    """
    saida = []
    for amp in resultado.amostras:
        ano = amp.ano_verdade
        if ano is None and ano_padrao is not None:
            ano = ano_padrao
        
        saida.append({
            "prefixo": amp.prefixo,
            "numero_sequencial": amp.numero_sequencial,
            "ano_verdade": ano,
            "cts": amp.cts.copy(),
        })
    return saida


# --------------------------------------------------------------------------- #
# CLI para teste rápido                                                          #
# --------------------------------------------------------------------------- #

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 3:
        print("Uso: python -m src.parser_termociclador <arquivo_1_4.xlsx> <arquivo_2_3.xlsx>")
        sys.exit(1)
    
    arquivo_1_4 = Path(sys.argv[1])
    arquivo_2_3 = Path(sys.argv[2])
    
    with open(arquivo_1_4, "rb") as f:
        r1 = parse_arquivo_termociclador(f.read(), arquivo_1_4.name)
    
    with open(arquivo_2_3, "rb") as f:
        r2 = parse_arquivo_termociclador(f.read(), arquivo_2_3.name)
    
    merged = merge_arquivos_termociclador(r1, r2)
    
    print(f"=== Parse Dengue 1-4 ({arquivo_1_4.name}) ===")
    print(f"Amostras: {len(r1.amostras)}")
    print(f"Sem ano: {r1.sample_ids_sem_ano}")
    print(f"Erros: {r1.erros}")
    
    print(f"\n=== Parse Dengue 2-3 ({arquivo_2_3.name}) ===")
    print(f"Amostras: {len(r2.amostras)}")
    print(f"Sem ano: {r2.sample_ids_sem_ano}")
    print(f"Erros: {r2.erros}")
    
    print(f"\n=== Merge ===")
    print(f"Amostras totais: {len(merged.amostras)}")
    print(f"Sem ano: {merged.sample_ids_sem_ano}")
    print(f"Erros: {merged.erros}")
    
    # Mostra primeiras 5 amostras
    for amp in merged.amostras[:5]:
        cts_str = ", ".join(f"{k}={v}" for k, v in amp.cts.items() if v is not None)
        ano_str = f"/{amp.ano_verdade % 100:02d}" if amp.ano_verdade else " (sem ano)"
        print(f"  {amp.prefixo}{amp.numero_sequencial}{ano_str}: {cts_str}")
