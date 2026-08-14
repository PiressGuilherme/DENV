"""Geração do mapa de trabalho para extração de Dengue.

O arquivo é produzido a partir de um modelo versionado. Apenas as células de
entrada são alteradas; fórmulas, estilos, mesclagens, validações e configuração
de impressão permanecem sob responsabilidade do modelo.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


RAIZ = Path(__file__).resolve().parent.parent
MODELO_MAPA_EXTRACAO = RAIZ / "data" / "modelos" / "mapa_extracao.xlsx"

ANO_TRABALHO = 2026
MAX_AMOSTRAS = 94
POSICOES_AMOSTRAS: tuple[str, ...] = tuple(
    f"{coluna}{linha}"
    for coluna in "ABCDEFGHIJKL"
    for linha in range(2, 10)
)[:MAX_AMOSTRAS]
POSICAO_CN = "L8"
POSICAO_CP = "L9"

OPERADORES: tuple[str, ...] = (
    "Anna",
    "Daniel",
    "Daniela",
    "Guilherme",
    "João",
    "Juliano",
)


class ErroMapaExtracao(ValueError):
    """Seleção ou metadado incompatível com o mapa de extração."""


@dataclass(frozen=True)
class AmostraExtracao:
    chave: str
    numero_sequencial: int
    ano_verdade: int


@dataclass(frozen=True)
class MapaExtracao:
    conteudo: bytes
    nome_arquivo: str
    ensaio: str
    quantidade_amostras: int


def validar_amostras(amostras: Iterable[AmostraExtracao]) -> list[AmostraExtracao]:
    selecionadas = list(amostras)
    if not selecionadas:
        raise ErroMapaExtracao("Selecione ao menos uma amostra.")
    if len(selecionadas) > MAX_AMOSTRAS:
        raise ErroMapaExtracao(
            f"Uma placa comporta no máximo {MAX_AMOSTRAS} amostras, além de CN e CP."
        )

    ids_logicos: list[tuple[int, int]] = []
    for amostra in selecionadas:
        chave = amostra.chave.strip().upper()
        match = re.fullmatch(r"D(\d+)/(\d{2}|\d{4})", chave)
        if not match:
            raise ErroMapaExtracao(
                f"Amostra {amostra.chave!r} não possui um Número Interno Dengue válido."
            )
        if amostra.ano_verdade != ANO_TRABALHO:
            raise ErroMapaExtracao(
                f"Amostra {amostra.chave} pertence a {amostra.ano_verdade}; "
                f"este mapa aceita somente amostras de {ANO_TRABALHO}."
            )
        if isinstance(amostra.numero_sequencial, bool):
            raise ErroMapaExtracao(f"Número inválido em {amostra.chave}.")
        try:
            numero = int(amostra.numero_sequencial)
        except (TypeError, ValueError):
            raise ErroMapaExtracao(f"Número inválido em {amostra.chave}.") from None
        if numero <= 0 or numero != amostra.numero_sequencial:
            raise ErroMapaExtracao(f"Número inválido em {amostra.chave}.")

        numero_chave = int(match.group(1))
        ano_chave_raw = match.group(2)
        ano_chave = int(ano_chave_raw)
        if len(ano_chave_raw) == 2:
            ano_chave += 2000
        if numero_chave != numero or ano_chave != amostra.ano_verdade:
            raise ErroMapaExtracao(
                f"Dados inconsistentes para a amostra {amostra.chave}."
            )
        ids_logicos.append((numero, amostra.ano_verdade))

    if len(set(ids_logicos)) != len(ids_logicos):
        raise ErroMapaExtracao("A seleção contém amostras duplicadas.")

    return sorted(selecionadas, key=lambda amostra: amostra.numero_sequencial)


def formatar_numero_interno(amostra: AmostraExtracao) -> str:
    """Formata ``D447/26`` como ``447/26``, sem zero artificial à esquerda."""
    return f"{int(amostra.numero_sequencial)}/{amostra.ano_verdade % 100:02d}"


def montar_ensaio(data_extracao: date, numero_placa: int) -> str:
    if not isinstance(data_extracao, date):
        raise ErroMapaExtracao("Data da extração inválida.")
    if isinstance(numero_placa, bool):
        raise ErroMapaExtracao("Número da placa deve ser um inteiro positivo.")
    try:
        placa = int(numero_placa)
    except (TypeError, ValueError):
        raise ErroMapaExtracao("Número da placa deve ser um inteiro positivo.") from None
    if placa <= 0 or placa != numero_placa:
        raise ErroMapaExtracao("Número da placa deve ser um inteiro positivo.")
    return f"DENV{data_extracao:%d%m%y}-{placa}"


def gerar_mapa_extracao(
    amostras: Iterable[AmostraExtracao],
    *,
    data_extracao: date,
    numero_placa: int = 1,
    operador: Optional[str] = None,
    caminho_modelo: Path = MODELO_MAPA_EXTRACAO,
) -> MapaExtracao:
    """Preenche o modelo e retorna um XLSX pronto para download."""
    selecionadas = validar_amostras(amostras)
    ensaio = montar_ensaio(data_extracao, numero_placa)

    operador_normalizado = (operador or "").strip()
    if operador_normalizado and operador_normalizado not in OPERADORES:
        raise ErroMapaExtracao("Operador inválido para o modelo de extração.")
    if not caminho_modelo.is_file():
        raise ErroMapaExtracao("Modelo da planilha de extração não encontrado.")

    # rich_text=True preserva os diferentes tamanhos de fonte dentro de uma
    # mesma célula (por exemplo, D2 e M1 da aba Extração).
    workbook = load_workbook(caminho_modelo, data_only=False, rich_text=True)
    try:
        if not {"Amostras", "Extração"}.issubset(workbook.sheetnames):
            raise ErroMapaExtracao("Modelo da extração não contém as abas esperadas.")

        aba_amostras = workbook["Amostras"]
        aba_extracao = workbook["Extração"]

        for posicao in POSICOES_AMOSTRAS:
            aba_amostras[posicao] = None
        for posicao, amostra in zip(POSICOES_AMOSTRAS, selecionadas):
            aba_amostras[posicao] = formatar_numero_interno(amostra)

        aba_amostras[POSICAO_CN] = "CN"
        aba_amostras[POSICAO_CP] = "CP"
        aba_amostras["B1"] = ensaio
        aba_extracao["K16"] = operador_normalizado or None

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

        buffer = BytesIO()
        workbook.save(buffer)
    finally:
        workbook.close()

    nome_arquivo = (
        f"DENGUE - {int(numero_placa)} - MAPA DE TRABALHO PARA EXTRAÇÃO "
        f"{data_extracao:%d_%m_%Y}.xlsx"
    )
    return MapaExtracao(
        conteudo=buffer.getvalue(),
        nome_arquivo=nome_arquivo,
        ensaio=ensaio,
        quantidade_amostras=len(selecionadas),
    )
